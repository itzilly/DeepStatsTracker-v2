from openpyxl import Workbook, load_workbook
from bin.scripts import getLogger


def create_workbook(uuid, player_name):
    path = 'bin/PlayerData/' + uuid + '.xlsx'
    getLogger.info(f"Creating workbook for {player_name}: {path}")
    individual_titles = ['Date', 'Final Kills', 'Final Void Kills', 'Final Projectile Kills',
                         'Final Fall Kills', 'Final Explosion Kills', 'Final Magic Kills',
                         'Final Suffocation Kills', 'Final Deaths', 'Final Void Deaths',
                         'Final Projectile Deaths', 'Final Fall Deaths', 'Final Explosion Deaths',
                         'Final Magic Deaths', 'Final Suffocation Deaths', 'FKDR', 'Kills',
                         'Void Kills', 'Projectile Kills', 'Fall Kills', 'Explosion Kills', 'Magic Kills',
                         'Suffocation Kills', 'Deaths', 'Void Deaths', 'Projectile Deaths',
                         'Fall Deaths', 'Explosion Deaths', 'Magic Deaths', 'Suffocation Deaths',
                         'KDR', 'Games Played', 'Wins', 'Losses', 'WLR', 'Iron', 'Gold', 'Diamonds',
                         'Emeralds', 'Total Resources', 'Purchases', 'Permanent Purchases']

    workbook = Workbook()

    sheet = workbook.active

    overall_sheet = workbook.create_sheet("Bedwars Overall")
    overall_sheet.title = "Bedwars Overall"
    overall_sheet = workbook.get_sheet_by_name("Bedwars Overall")
    overall_titles = ['Date', 'Stars', 'Experience', 'Final Kills', 'Final Deaths', 'FKDR', 'Kills', 'Deaths', 'KDR', 'Beds Broken', 'Beds Lost', 'BBLR', 'Games Played', 'Wins', 'Losses', 'WLR', 'Coins', 'Loot Boxes', 'Iron Collected', 'Gold Collected', 'Diamonds Collected', 'Emeralds Collected']
    overall_sheet.append(overall_titles)

    solo_sheet = workbook.create_sheet("Bedwars Solos")
    solo_sheet.title = "Bedwars Solos"
    solo_sheet = workbook.get_sheet_by_name("Bedwars Solos")

    normals_titles = [
        'Date', 'Stars', 'Final Kills', 'Final Void Kills', 'Final Kills Non Void', 'Final Projectile Kills',
        'Final Fall Kills', 'Final Explosion Kills', 'Final Magic Kills', 'Final Suffocation Kills',
        'Final Deaths', 'Final Void Deaths', 'Final Deaths Non World', 'Final Projectile Deaths',
        'Final Fall Deaths', 'Final Explosions Deaths', 'Final Magic Deaths', 'Final Suffocation Deaths',
        'Kills', 'Void Kills', 'Kills Non Void', 'Projectile Kills', 'Fall Kills', 'Explosion Kills',
        'Magic Kills', 'Suffocation Kills', 'deaths', 'Void Deaths', 'Deaths Non Void', 'Projectile Deaths',
        'Fall Deaths', 'Explosion Deaths', 'Magic Deaths', 'Suffocation Deaths', 'FKDR', 'KDR', 'Games Played',
        'Wins', 'Losses', 'Iron', 'Gold', 'Diamonds', 'Emeralds', 'Total Resources', 'Iron Per Game', 'Gold Per Game',
        'Diamonds Per Game', 'Emeralds Per Game', 'Total Resources Per Game', 'Purchases', 'Permanent Purchases',
        'Beds Broken', 'Beds Lost', 'BBLR', 'Beds Broken Per Game'
    ]

    solo_sheet.append(normals_titles)

    duo_sheet = workbook.create_sheet("Bedwars Duos")
    duo_sheet.title = "Bedwars Duos"
    duo_sheet = workbook.get_sheet_by_name("Bedwars Duos")
    duo_sheet.append(normals_titles)

    threes_sheet = workbook.create_sheet("Bedwars Threes")
    threes_sheet.title = "Bedwars Threes"
    threes_sheet = workbook.get_sheet_by_name("Bedwars Threes")
    threes_sheet.append(normals_titles)

    fours_sheet = workbook.create_sheet("Bedwars Fours")
    fours_sheet.title = "Bedwars Fours"
    fours_sheet = workbook.get_sheet_by_name("Bedwars Fours")
    fours_sheet.append(normals_titles)
    
    FoF_sheet = workbook.create_sheet("Bedwars FoF")
    FoF_sheet.title = "Bedwars FoF"
    FoF_sheet = workbook.get_sheet_by_name("Bedwars FoF")
    FoF_sheet.append(normals_titles)

    workbook.remove_sheet(sheet)
    workbook.save(path)

    getLogger.info(f"Created and saved workbook for {player_name}: \'{uuid}\'")


def save_overall(uuid, stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    overall_sheet = workbook["Bedwars Overall"]
    overall_sheet.append(stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS OVERALL")


def save_solos(uuid, solo_stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    solos_sheet = workbook["Bedwars Solos"]
    solos_sheet.append(solo_stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS SOLOS")


def save_duos(uuid, solo_stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    solos_sheet = workbook["Bedwars Duos"]
    solos_sheet.append(solo_stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS DUOS")


def save_threes(uuid, solo_stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    solos_sheet = workbook["Bedwars Threes"]
    solos_sheet.append(solo_stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS SOLOS")


def save_fours(uuid, solo_stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    solos_sheet = workbook["Bedwars Fours"]
    solos_sheet.append(solo_stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS Fours")


def save_FoF(uuid, solo_stats):
    path = 'bin/playerData/' + uuid + '.xlsx'
    workbook = load_workbook(path)
    solos_sheet = workbook["Bedwars FoF"]
    solos_sheet.append(solo_stats)
    workbook.save(path)

    getLogger.info("Saving player data to " + path + " BEDWARS FoF")
